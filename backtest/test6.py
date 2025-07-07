import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy


def copy_sheet_with_format(source_file, target_file, sheet_name):
    # 打开源文件和目标文件
    source_wb = openpyxl.load_workbook(source_file)
    target_wb = openpyxl.load_workbook(target_file)

    # 检查源文件是否包含该工作表
    if sheet_name not in source_wb.sheetnames:
        print(f"源文件中未找到工作表: {sheet_name}")
        return

    source_sheet = source_wb[sheet_name]

    # 如果目标文件已有该工作表，则先删除
    if sheet_name in target_wb.sheetnames:
        target_wb.remove(target_wb[sheet_name])

    # 创建一个新的工作表
    new_sheet = target_wb.create_sheet(title=sheet_name)

    # 复制单元格内容、公式和格式
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet[cell.coordinate]
            new_cell.value = cell.value  # 复制值（包括公式）

            # 复制格式（字体、填充、对齐方式、边框等）
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 复制列宽
    for col_idx, col_dim in source_sheet.column_dimensions.items():
        new_sheet.column_dimensions[col_idx].width = col_dim.width

    # 复制行高
    for row_idx, row_dim in source_sheet.row_dimensions.items():
        new_sheet.row_dimensions[row_idx].height = row_dim.height

    # 复制合并单元格
    for merged_range in source_sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_range))

    # 保存目标文件
    target_wb.save(target_file)
    print(f"工作表 '{sheet_name}' 已成功复制并替换到 {target_file}")


file_name_list=["北仑工业","北仑住宅","北仑商业","奉化商业","奉化住宅","奉化工业","高新工业","高新商业","高新住宅","海曙工业","海曙商业","海曙住宅","江北工业","江北商业","江北住宅","鄞州工业","鄞州商业","鄞州住宅","镇海工业","镇海商业","镇海住宅",]

sheet_name_list = ["土地现场查勘表", "基础资料", "建筑面积等"]
for filename in file_name_list:
    source_file = "C:\\Users\yizer\Desktop\\4、标定地价测算表\\4、标定地价测算表\\"+filename+".xlsx"
    target_file = "C:\\Users\yizer\Desktop\\4、标定地价测算表\\4、标定地价测算表\\2025\新建文件夹\\"+filename+".xlsx"
    for sheet_name in sheet_name_list:
        copy_sheet_with_format(source_file, target_file, sheet_name)
