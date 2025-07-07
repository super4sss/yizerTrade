import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime


def transfer_excel_data(data_file, template_file, output_dir):
    """
    将数据文件中的每一行数据填充到模板文件中，并生成多个填充后的Excel文件

    参数:
    data_file (str): 数据源Excel文件路径
    template_file (str): 模板Excel文件路径
    output_dir (str): 输出目录路径
    """
    # 创建输出目录（如果不存在）
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 读取数据源Excel文件，明确指定空值处理
    df = pd.read_excel(data_file, keep_default_na=True)

    # 预处理数据：确保真正的空值为None，而不是空字符串、NaN等
    for column in df.columns:
        # 将NaN、空字符串等统一处理为None
        df[column] = df[column].apply(lambda x: None if pd.isna(x) or (isinstance(x, str) and x.strip() == '') else x)
    print(f"共读取 {len(df)} 条数据记录")

    # 定义数据源列名与模板中占位符的映射关系
    field_mapping = {
        '交易点编码': '#(交易点编码)',
        '宗地编号': '#(宗地编号)',
        '所在城市': '#(所在城市)',
        '宗地名称': '#(宗地名称)',
        '宗地坐落': '#(宗地坐落)',
        '所在土地一级类': '#(所在土地一级类)',
        '所在土地二级类': '#(所在土地二级类)',
        '所在土地三级类': '#(所在土地三级类)',
        '所在土地级别': '#(所在土地级别)',
        '旧区段编码': '#(旧区段编码)',
        '所在区段编码': '#(所在区段编码)',
        '受让单位': '#(受让单位)',
        '出让用途': '#(出让用途)',
        '主要出让用途': '#(主要出让用途)',
        '出让开发程度': '#(出让开发程度)',
        '建筑密度': '#(建筑密度)',
        '绿化率': '#(绿化率)',
        '宗地面积平方米': '#(宗地面积)',
        '容积率': '#(容积率)',
        '容积率最大值': '#(容积率最大值)',
        '成交时间': '#(成交时间)',
        '出让年限': '#(出让年限)',
        '出让方式': '#(出让方式)',
        '成交公示号': '#(成交公示号)',
        '成交总价万元': '#(成交总价万元)',
        '成交楼面价': '#(成交楼面价元平方米)',
        '成交地面价': '#(成交地面价元平方米)',
        '规划建筑面积平方米': '#(规划建筑面积平方米)',
        '住宅剥离价格': '#(住宅剥离价格)',
        '建筑限高': '#(建筑限高)',
        '用途剥离': '#(用途剥离)',
        '配建剥离': '#(配建剥离)',
        '产权限制修正': '#(产权限制修正)',
        '用途剥离内容': '#(用途剥离内容)',
        '配建内容': '#(配建内容)',
        '具体产权限制内容': '#(具体产权限制内容)',
        '其他修正': '#(其他修正)',
        '产权限制修正系数': '#(产权限制修正系数)',
        '容积率修正': '#(容积率修正)',
        '交易点价格修正后地面价元平方米': '#(修正后地面价元平方米)',
        '交易点价格修正后楼面价元平方米': '#(修正后楼面价元平方米)',
        '设定容积率': '#(设定容积率)',
    }

    # 反向查找映射（从占位符到列名）
    reverse_mapping = {v: k for k, v in field_mapping.items()}

    # 记录成功和失败的数量
    success_count = 0
    error_count = 0

    # 处理每一行数据
    for index, row in df.iterrows():
        try:
            # 加载模板文件（为每一行数据创建一个新的工作簿实例）
            wb = load_workbook(template_file)
            ws = wb.active
            # 查找并替换模板中的占位符
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    cell_value = cell.value
                    # print(cell_value)
                    # 检查单元格是否包含占位符
                    if isinstance(cell_value, str) and cell_value.startswith('#') and cell_value in reverse_mapping:
                        field_name = reverse_mapping[cell_value]
                        if field_name in row:
                            print(row['所在区段编码'])
                            # 替换占位符为实际数据
                            cell.value = row[field_name]

            # 填写当前日期到"填表日期"单元格
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value == "填表日期":
                        # 在"填表日期"右侧单元格填写当前日期
                        date_cell = ws.cell(row=r, column=c + 1)
                        date_cell.value = datetime.now().strftime("%Y/%m/%d")

            # 填写填表人信息（如果有的话）
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value == "填表单位":
                        # 在"填表单位"右侧单元格填写信息
                        unit_cell = ws.cell(row=r, column=c + 1)
                        unit_cell.value = "数据导入系统"

            # 保存填充后的模板文件
            record_id = str(row.get('交易点编码', f'记录{index + 1}'))
            output_file = os.path.join(output_dir, f"宁波市交易点登记表({record_id}).xlsx")
            wb.save(output_file)
            print(f"已生成文件: {output_file}")
            success_count += 1

        except Exception as e:
            print(f"处理记录 {index + 1} 时出错: {str(e)}")
            error_count += 1

    print(f"处理完成。成功: {success_count}, 失败: {error_count}, 总记录数: {len(df)}")


if __name__ == "__main__":
    # 定义文件路径
    data_file = "F:\\副本2025年3-2025年5月交易样点汇总表1.xlsx"  # 数据源文件路径（图2所示文件）
    template_file = "F:\\交易点登记表 (2).xlsx"  # 模板文件路径（图1所示文件）
    output_dir = "F:\\填充后表格"  # 输出目录

    # 执行数据转移
    transfer_excel_data(data_file, template_file, output_dir)
